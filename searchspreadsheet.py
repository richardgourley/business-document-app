import os
import openpyxl
from openpyxl.utils import get_column_letter

class SearchSpreadsheet:
    def __init__(self):
        self.check_excel_files_exist()
        self.display_available_files()
        self.choose_excel_file()
        self.get_column_names()
        self.choose_search_term()
        self.print_search_results()

    def check_excel_files_exist(self):
        os.chdir('excelfiles')
        self.excel_files = os.listdir()
        if len(self.excel_files) == 0:
            print('============')
            print("We couldn't find any files in the 'excelfiles' directory. Please add some files and run the program again.")
            print('============')
            quit()

    def display_available_files(self):
        print('Here are the available excel files you can search.')
        print('===============')
        for file in self.excel_files:
            print(file)
        print('===============')

    def choose_excel_file(self):
        excel_file = None
        while excel_file is None:
            print("Enter the name of a file to search. (Must match a file name shown above)")
            excel_file = input()
            try:
                excel_file = openpyxl.load_workbook(excel_file)
            except:
                excel_file = None
                print("Sorry, the file name you entered doesn't match an available file.")
        
        self.excel_file = excel_file
        self.sheet = self.excel_file.active
        print("===============")

    def get_column_names(self):
        self.column_names = ""
        for i in range(1, self.sheet.max_column + 1):
            self.column_names += str(self.sheet[get_column_letter(i) + "1"].value) + " -- "

    def choose_search_term(self):
        search_term = ""
        while search_term == "":
            print("Enter a search term. You can enter words or numbers. (Can't be blank)")
            search_term = input()
        self.search_term = search_term

        print("=============")

    def print_search_results(self):
        number_of_results = 0

        search_term_lower = self.search_term.lower()

        print(self.column_names)

        for row in list(self.sheet.rows):
            for cell in row:
                if search_term_lower in str(cell.value).lower():
                    number_of_results += 1
                    print(", ".join(map(lambda x: str(x.value), row)))
                    break
        print("TOTAL RESULTS = ", number_of_results)
        