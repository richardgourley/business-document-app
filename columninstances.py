import os
import openpyxl
from openpyxl.utils import get_column_letter

class CountColumnInstances:
    def __init__(self):
        self.check_excel_files_exist()
        self.display_available_files()
        self.choose_excel_file_and_assign_class_variables()
        #self.display_columns()
        #self.choose_a_column()
        #self.print_results()

    def check_excel_files_exist(self):
        os.chdir('excelfiles')
        self.excel_files = os.listdir()
        if len(self.excel_files) == 0:
            print('============')
            print("We couldn't find any files in the 'excelfiles' directory. Please add some files and run the program again.")
            print('============')
            quit()

    def display_available_files(self):
        print('Here are the available excel files you can search.')
        print('===============')
        for file in self.excel_files:
            print(file)
        print('===============')

    def choose_excel_file_and_assign_class_variables(self):
        excel_file = None
        while excel_file is None:
            print("Enter the name of a file to search. (Must match a file name shown above)")
            excel_file = input()
            try:
                excel_file = openpyxl.load_workbook(excel_file)
            except:
                excel_file = None
                print("Sorry, the file name you entered doesn't match an available file.")
        
        self.assign_class_variables(excel_file)
        print("===============")

    def assign_class_variables(self, excel_file):
        self.excel_file = excel_file
        self.sheet = excel_file.active
        self.sheet_column_letters = list()
        for i in range(1, self.sheet.max_column + 1):
            self.sheet_column_letters.append(get_column_letter(i))

    def display_columns(self):
        print('Here are the available columns.')
        print('===============')
        # Will use this list in CHOOSE_A_COLUMN method
        self.sheet_column_letters = list()

        for i in range(1, self.sheet.max_column + 1):
            print(i, "=", get_column_letter(i), self.sheet[get_column_letter(i) + "1"].value)
            self.sheet_column_letters.append(get_column_letter(i))
        print('===============')

    def choose_a_column(self):
        chosen_column_number = None

        while chosen_column_number == None:
            print("Enter a column number from above")
            chosen_column_number = input()

            for i in range(1, self.sheet.max_column + 1):
                if str(i) == chosen_column_number:
                    self.chosen_column_number = chosen_column_number
                    break
            
            try:
                print("Chosen column number is: ", str(self.chosen_column_number))
            except:
                print("Sorry, number not valid")
                chosen_column_number = None

        print("=============")

    def print_results(self):
        results = {}
        count = (self.sheet.max_row -1)

        print("From " + str(count) + " rows, we found:")

        for i in range(2, self.sheet.max_row + 1):
            # get value
            value = self.sheet[self.column_letter + str(i)].value
            # see if value is a key in results dictionary
            # if not, create a new key set to 1, if yes, add 1 to the value of the key
            if not value in results:
                results[value] = 1
            else:
                results[value] += 1

        # 'results' sorted by highest number of instances first
        # 'results_sorted' returns list of tuples
        results_sorted = sorted( [ (v,k) for k,v in results.items() ], reverse=True )

        # print no. of instances and column name
        print("NUM INSTANCES,", self.sheet[self.column_letter + "1"].value)
        # loop the sorted results 
        for result in results_sorted:
            # create a string to display each item in the result
            result_row = ""
            for item in result:
                result_row += str(item) + " "
            print(result_row)






